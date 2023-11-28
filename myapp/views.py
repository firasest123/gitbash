from .models import case, VisitedLink, CopyMacro, UserLoginLogoutRecord, StatusChange, Campagne, Winback2, Winback2Results, CampagneResults,Ningen_Staff,Activite,Ningen_Management, upload
from django.conf import settings
import os
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.csrf import csrf_exempt
import json
from .forms import VisitedLinkForm, DateRangeForm, ChangePasswordForm, HorairePreferenceForm, CampagneForm,Winback2ResultsForm,NingenStaffForm,ActiviteForm,NingenManagementForm
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import AuthenticationForm
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from datetime import timedelta, datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from django.db import connection
from django.contrib.auth.decorators import login_required
import pandas as pd
from django.urls import reverse
import csv
@login_required(login_url='user_login')
def index(request):
    return render(request, 'index.html')


@login_required(login_url='user_login')
def jumia_ods_tn(request):
    return render(request, 'jumia_ods_tn.html')


@login_required(login_url='user_login')
def get_case_suggestions(request):
    query = request.GET.get('q')
    suggestions = []

    if query:
        cases = case.objects.filter(numero__icontains=query)
        suggestions = [{'numero': case.numero, 'nom': case.nom} for case in cases]

    return JsonResponse({'suggestions': suggestions})

@login_required(login_url='user_login')
def case_detail(request, case_numero, case_nom):
    case_obj = get_object_or_404(case, numero=case_numero, nom=case_nom)
    return render(request, 'case_detail.html', {'case_obj': case_obj})

def serve_image(request, image_name):
    image_path = os.path.join(settings.MEDIA_ROOT, 'images', image_name)
    with open(image_path, 'rb') as img_file:
        return HttpResponse(img_file.read(), content_type='image/png')


@login_required(login_url='user_login')
def case_list_view(request, case_name):
    cases = case.objects.filter(nom=case_name)
    return render(request, 'case_list.html', {'cases': cases, 'case_name': case_name})

@login_required(login_url='user_login')
def commande_annulee_view(request):
    return render(request, 'commande_annulee.html')

def custom_login(request):

    if request.user.is_authenticated:
        return redirect('index')

    if request.method == "POST":
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = authenticate(
                request,
                username=form.cleaned_data["username"],
                password=form.cleaned_data["password"],
            )
            if user is not None:
                login(request, user)


                user_agent_string = request.META.get('HTTP_USER_AGENT', '')


                if 'Mobile' in user_agent_string:
                    device_type = 'Mobile'
                elif 'Tablet' in user_agent_string:
                    device_type = 'Tablet'
                else:
                    device_type = 'Computer'

                login_record = UserLoginLogoutRecord(
                    user=user,
                    login_time=timezone.now(),
                    device=device_type,
                )
                login_record.save()

                return redirect("index")
    else:
        form = AuthenticationForm()
    return render(request, "registration/login.html", {"form": form})


def custom_logout(request):
    if request.user.is_authenticated:
        try:
            last_login_record = UserLoginLogoutRecord.objects.filter(
                user=request.user,
                logout_time=None
            ).latest('login_time')
            last_login_record.logout_time = timezone.now()

            open_status_change = StatusChange.objects.filter(
                user=request.user,
                end_time=None
            ).first()

            if open_status_change:
                open_status_change.end_time = last_login_record.logout_time
                open_status_change.save()
            last_login_record.save()
        except UserLoginLogoutRecord.DoesNotExist:
            pass
        logout(request)
    return redirect('user_login')




def download_interactions(request):
    if not request.user.is_superuser:
        return HttpResponse("You don't have permission to access this page.")

    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    if not start_date_str or not end_date_str:
        return HttpResponse("Please provide both start date and end date.")

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        return HttpResponse("Invalid date format. Please use YYYY-MM-DD.")

    interactions = VisitedLink.objects.filter(
        date__range=(start_date, end_date)
    ).values('user__name', 'url', 'date', 'time')

    df = pd.DataFrame.from_records(interactions)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=interactions.xlsx'

    df.to_excel(response, index=False, sheet_name='Interactions')

    return response


@csrf_exempt
def save_copy_macro(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        text = data.get('text')
        link = data.get('link')
        user = request.user

        if text and link and user.is_authenticated:
            CopyMacro.objects.create(user=user, text=text, link=link)

        return HttpResponse(status=200)
    return HttpResponse(status=400)

def download_copied_macros(request):
    if not request.user.is_authenticated or not request.user.is_staff:
        return HttpResponse("You don't have permission to access this page.")

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if not start_date or not end_date:
        return HttpResponse("Please provide both start date and end date.")

    start_date = datetime.strptime(start_date, "%Y-%m-%d")
    end_date = datetime.strptime(end_date, "%Y-%m-%d")

    copied_macros = CopyMacro.objects.filter(
        date__range=(start_date, end_date)
    ).values('user__name', 'text', 'link', 'date', 'time')

    df = pd.DataFrame(copied_macros)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=copied_macros.xlsx'
    df.to_excel(response, index=False)

    return response

@login_required(login_url='user_login')
def download_user_login_logout(request):
    if not request.user.is_authenticated or not request.user.is_staff:
        return HttpResponse("You don't have permission to access this page.")

    if request.method == "GET":
        form = DateRangeForm(request.GET)

        if form.is_valid():
            start_date = form.cleaned_data["start_date"]
            end_date = form.cleaned_data["end_date"]

            user_login_logout_entries = UserLoginLogoutRecord.objects.filter(
                login_time__date__range=(start_date, end_date)
            ).values(
                "user__name",
                "login_time__date",
                "login_time__time",
                "logout_time__date",
                "logout_time__time",
                "device",
            )

            if not user_login_logout_entries:
                user_login_logout_entries = [
                    {
                        "user__name": "No records found",
                        "login_time__date": "",
                        "login_time__time": "",
                        "logout_time__date": "",
                        "logout_time__time": "",
                        "device": "",
                        "duration": "",
                    }
                ]

            user_login_logout_entries = list(user_login_logout_entries)

            for entry in user_login_logout_entries:
                entry["login_date"] = entry.pop("login_time__date")
                entry["login_time"] = entry["login_time__time"].strftime('%H:%M') if entry[
                    "login_time__time"] else ""
                entry["logout_date"] = entry.pop("logout_time__date")
                entry["logout_time"] = entry["logout_time__time"].strftime('%H:%M') if entry[
                    "logout_time__time"] else ""

                login_time = entry.get("login_time__time")
                logout_time = entry.get("logout_time__time")

                if login_time and logout_time:
                    try:
                        login_datetime = datetime.combine(datetime.today(), login_time)
                        logout_datetime = datetime.combine(datetime.today(), logout_time)

                        duration = logout_datetime - login_datetime

                        hours, remainder = divmod(duration.seconds, 3600)
                        minutes, seconds = divmod(remainder, 60)
                        entry["duration"] = f"{hours:02}:{minutes:02}:{seconds:02}"
                    except Exception as e:
                        entry["duration"] = ""  
                else:
                    entry["duration"] = ""

                
                del entry["login_time__time"]
                del entry["logout_time__time"]

            df = pd.DataFrame(user_login_logout_entries)
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = f"attachment; filename=user_login_logout_data.xlsx"
            df.to_excel(response, index=False)
            return response
    else:
        form = DateRangeForm()

    return render(request, "index.html", {"form": form})


def change_password(request):
    if request.method == 'POST':
        form = ChangePasswordForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            ancien_mot_de_passe = form.cleaned_data['ancien_mot_de_passe']
            nouveau_mot_de_passe_1 = form.cleaned_data['nouveau_mot_de_passe_1']
            nouveau_mot_de_passe_2 = form.cleaned_data['nouveau_mot_de_passe_2']

            user = authenticate(request, username=username, password=ancien_mot_de_passe)

            if user is not None:
                if nouveau_mot_de_passe_1 == nouveau_mot_de_passe_2:
                    user.set_password(nouveau_mot_de_passe_1)
                    user.save()

                    login(request, user)

                    return redirect('index')
                else:
                    form.add_error('nouveau_mot_de_passe_2', 'Les nouveaux mots de passe ne correspondent pas.')
            else:
                form.add_error('ancien_mot_de_passe', 'Nom d''utilisateur ou ancien mot de passe invalide.')
    else:
        form = ChangePasswordForm()

    return render(request, 'change_password.html', {'form': form})


@login_required(login_url='user_login')
def change_status(request, status):
    if request.user.is_authenticated:
        user = request.user
        now = timezone.now()

        open_status_change = StatusChange.objects.filter(user=user, end_time=None).first()

        if open_status_change:
            open_status_change.end_time = now
            open_status_change.save()

        status_change = StatusChange(user=user, status=status, start_time=now)
        status_change.save()

    return redirect('index')


@login_required(login_url='user_login')
def download_status_excel(request):
    if not request.user.is_authenticated or not request.user.is_staff:
        return HttpResponse("You don't have permission to access this page.")

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if not start_date or not end_date:
        return HttpResponse("Please provide both start date and end date.")

    start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

    status_changes = StatusChange.objects.filter(
        start_time__date__range=[start_date, end_date],
        end_time__date__range=[start_date, end_date]
    ).values('user__name', 'status', 'start_time', 'end_time')

    wb = Workbook()
    ws = wb.active
    ws.title = "Status Changes"

    headers = ['User', 'Status', 'Start Date', 'Start Time', 'End Date', 'End Time', 'Duration']

    for col_num, header_title in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = ws['{}1'.format(col_letter)]
        cell.value = header_title
        cell.font = Font(bold=True)

    for row_num, status_change in enumerate(status_changes, 2):
        ws.cell(row=row_num, column=1, value=status_change['user__name'])
        ws.cell(row=row_num, column=2, value=status_change['status'])

        start_time = status_change['start_time'].replace(tzinfo=None)
        end_time = status_change['end_time'].replace(tzinfo=None)

        ws.cell(row=row_num, column=3, value=start_time.date())
        ws.cell(row=row_num, column=4, value=start_time.strftime('%H:%M'))
        ws.cell(row=row_num, column=5, value=end_time.date())
        ws.cell(row=row_num, column=6, value=end_time.strftime('%H:%M'))

        duration = end_time - start_time

        if isinstance(duration, timedelta):
            seconds = duration.total_seconds()
            hours, remainder = divmod(seconds, 3600)
            minutes, seconds = divmod(remainder, 60)
            duration = f"{int(hours):02}:{int(minutes):02}:{int(seconds):02}"
        else:
            duration = ""

        ws.cell(row=row_num, column=7, value=duration)

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=status_changes.xlsx'
    wb.save(response)

    return response


from django.shortcuts import render, redirect
import pandas as pd
from django.db import connection

def upload_excel(request):
    if request.method == "POST" and request.FILES["excel_file"]:
        excel_file = request.FILES["excel_file"]

        if not excel_file.name.endswith('.xlsx'):
            return render(request, "error.html", {"message": "Invalid file format. Please upload an Excel file."})

        df = pd.read_excel(excel_file)

        df = df.where(pd.notna(df), None)

        with connection.cursor() as cursor:
            for index, row in df.iterrows():
                gsm = row["GSM"]
                voucher_eligibilty = row["Voucher_eligibilty"]
                email = row["Email"]
                customer_id = row["Customer_ID"]
                voucher_code = row["Voucher code"]
                agent = row["Agent"]

                insert_sql = """
                    INSERT INTO besidedb.myapp_campagne (GSM, Voucher_eligibilty, Email, Customer_ID, Voucher_code, Agent)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """
                cursor.execute(insert_sql, [gsm, voucher_eligibilty, email, customer_id, voucher_code, agent])

        return redirect("index")

    return render(request, "upload_excel.html")



def agent_gsm_list(request):
    user = request.user

    if user.name in ['Samar Saadi', 'Riadh Mansouri', 'Ines Belghith', 'Chaima b.', 'khaoula w.'] or user.is_superuser:
        gsm_list = Campagne.objects.filter(Agent="nan", Done=False)
    else:
        gsm_list = Campagne.objects.filter(Agent=user, Done=False)

    campagne_list = Campagne.objects.filter(Done=False)
    campagne_results_done = CampagneResults.objects.filter(done=True).values_list('GSM', flat=True)
    campagne_list = campagne_list.exclude(GSM__in=campagne_results_done)

    return render(request, 'agent_gsm_list.html', {'gsm_list': gsm_list, 'campagne_list': campagne_list})


def agent_processed_forbidden_view(request):

    return render(request, 'agent_processed_forbidden.html')

def gsm_detail(request, gsm_id):
    gsm = get_object_or_404(Campagne, GSM=gsm_id)
    user = request.user



    if gsm.is_locked and gsm.locked_by != user:
        return redirect("agent_processed_forbidden")

    gsm.is_locked = True
    gsm.locked_by = user
    gsm.save()

    initial_data = {
        'Mail_client_Valide': gsm.Email,
    }
    form = CampagneForm(initial=initial_data)

    context = {
        'gsm_details': gsm,
        'form': form,
    }
    return render(request, 'gsm_detail.html', context)


from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.db import connection

def release_lock(request, gsm_id):
    if request.method == 'POST':
        gsm = get_object_or_404(Campagne, GSM=gsm_id)

        gsm.is_locked = False
        gsm.locked_by = None
        gsm.save()

        return JsonResponse({'success': True})

    return JsonResponse({'success': False})




from django.db import connection
from django.shortcuts import render, redirect
from django.utils import timezone

@login_required
def winback_form(request, gsm):
    if request.method == "POST":
        form = CampagneForm(request.POST)

        if form.is_valid():
            client_joignable = form.cleaned_data['Client_Joignable']
            raison_de_non_commande = form.cleaned_data['Raison_de_non_commande']
            feedback_post_voucher = form.cleaned_data['Feedback_post_Voucher']
            mail_client_valide = form.cleaned_data['Mail_client_Valide']
            commentaire = form.cleaned_data['commentaire']

            current_datetime = timezone.now().strftime("%Y-%m-%d %H:%M:%S")

            treated_by = request.user.name  

            with connection.cursor() as cursor:
                sql = """
                    UPDATE myapp_campagne
                    SET Done = TRUE,
                        Client_Joignable = %s,
                        Treated_By = %s,
                        Submission_DateTime = %s
                    WHERE GSM = %s
                """
                cursor.execute(sql, (client_joignable, treated_by, current_datetime, gsm))

            campagne_result = CampagneResults(
                treated_by=request.user,
                GSM=gsm,
                Client_Joignable=client_joignable,
                Raison_de_non_commande=raison_de_non_commande,
                Feedback_post_Voucher=feedback_post_voucher,
                Mail_client_Valide=mail_client_valide,
                commentaire=commentaire,
                done=True
            )
            campagne_result.save()

            return redirect("agent_gsm_list")
    else:
        form = CampagneForm()

    return render(request, 'gsm_detail.html')


@login_required(login_url="user_login")
def preference_regime_horaire_view(request):
    if request.method == 'POST':
        form = HorairePreferenceForm(request.POST)
        if form.is_valid():
            user = request.user
            new_table_entry = form.save(commit=False)
            new_table_entry.user = user
            new_table_entry.save()
            return redirect('index')
    else:
        form = HorairePreferenceForm()

    context = {
        'form': form,
    }
    return render(request, 'horaire_preference.html', context)


def release_lock2(request, gsm):
    winback2 = get_object_or_404(Winback2, GSM=gsm)

    if winback2.locked:
        winback2.locked = False
        winback2.save()
        return JsonResponse({'message': 'Lock released successfully'})
    else:
        return JsonResponse({'message': 'Page was not locked'})



def uploadwinback2(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']

        if excel_file.name.endswith(('.xls', '.xlsx')):
            df = pd.read_excel(excel_file)

            winback_instances = []

            for _, row in df.iterrows():
                winback_data = {
                    'Date_de_traitement': row['Date_de_traitement'],
                    'GSM': row['GSM'],
                    'VOUCHER_ELIGIBILITY': row['VOUCHER ELIGIBILTY'],
                    'CUSTOMER_ID': row['CUSTOMER ID'],
                    'CLIENT_JOIGNABLE': row['CLIENT JOIGNABLE'],
                    'RAISON_DE_NON_COMMANDE': row['RAISON DE NON COMMANDE'],
                    'FEEDBACK_POST_VOUCHER': row['FEEDBACK POST VOUCHER'],
                    'COMMENTAIRE': row['COMMENTAIRE'],
                    'Voucher_code': row['Voucher code'],
                    'Email': row['Email'],
                }
                winback_instance = Winback2(**winback_data)
                winback_instances.append(winback_instance)

            Winback2.objects.bulk_create(winback_instances)

            return redirect('index') 

    return render(request, 'uploadwinback2.html')


from django.shortcuts import render
from .models import Winback2, Winback2Results

from django.shortcuts import render
from django.db.models import Q
from .models import Winback2, Winback2Results


def winback_rappel_list(request):
    winback2_list = Winback2.objects.filter(Q(done=False))

    winback2_results_done = Winback2Results.objects.filter(done=True).values_list('GSM', flat=True)

    winback2_list = winback2_list.exclude(GSM__in=winback2_results_done)

    return render(request, 'WinBack_Rappel_list.html', {'winback2_list': winback2_list})




def winback2_detail(request, gsm):
    winback2 = get_object_or_404(Winback2, GSM=gsm)

    if not winback2.locked:
        winback2.locked = True
        winback2.save()

        return render(request, 'winback2_detail.html', {'winback2': winback2})
    else:
        return render(request, 'page_locked.html')


def winback_form_view(request, gsm):
    winback2 = get_object_or_404(Winback2, GSM=gsm)

    if request.method == 'POST':
        form = Winback2ResultsForm(request.POST)

        if form.is_bound:
            joignabilite = form.data['joignabilite']
            feedback_interet_voucher = form.data['feedback_interet_voucher']
            raison_non_interesse = form.data['raison_non_interesse']
            commentaire = form.data['commentaire']
            user = request.user
            winback2_results = Winback2Results(
                treated_by=user,
                GSM=gsm,
                joignabilite=joignabilite,
                feedback_interet_voucher=feedback_interet_voucher,
                raison_non_interesse=raison_non_interesse,
                commentaire=commentaire,
                done=True
            )
            winback2_results.save()

            winback2.done = True
            winback2.locked = True
            winback2.CLIENT_JOIGNABLE2 = joignabilite
            winback2.save()

            return redirect('winback_rappel_list') 
    else:
        form = Winback2ResultsForm()  

    return render(request, 'winback2_detail.html', {'form': form, 'winback2': winback2})


def create_staff(request):
    if request.method == 'POST':
        form = NingenStaffForm(request.POST)
        if form.is_valid():
            ningen_staff = form.save()
            id_beside = ningen_staff.ID_Beside
            return redirect('ajouter_activite', id_beside=id_beside)
    else:
        form = NingenStaffForm()

    return render(request, 'addStaff.html', {'form': form})



def ajouter_activite(request, id_beside):
    ningen_staff = Ningen_Staff.objects.get(ID_Beside=id_beside)

    if request.method == 'POST':
        form = ActiviteForm(request.POST)
        if form.is_valid():
            activite = form.save(commit=False)
            activite.agent = ningen_staff
            activite.save()

            if 'save_and_add' in request.POST:
                form = ActiviteForm()
            elif 'save' in request.POST:
                return redirect(reverse('staff_list'))
    else:
        form = ActiviteForm()

    return render(request, 'ajouter_activite.html', {'form': form, 'ningen_staff': ningen_staff})




from django.db.models import Q

def staff_list(request):
    query = request.GET.get('q', '')
    activite_filter = request.GET.get('activite_filter')
    micro_activite_filter = request.GET.get('micro_activite_filter')

    staff_members = Ningen_Staff.objects.all()

    if query:
        staff_members = staff_members.filter(
            Q(ID_Beside__icontains=query) |
            Q(Nom__icontains=query) |
            Q(Prenom__icontains=query)
        )

    if activite_filter:
        staff_members = staff_members.filter(Activité_actuelle=activite_filter)

    if micro_activite_filter:
        staff_members = staff_members.filter(Micro_activité_actuelle=micro_activite_filter)

    return render(request, 'staff_list.html', {
        'staff_members': staff_members,
        'query': query,
        'activite_filter': activite_filter,
        'micro_activite_filter': micro_activite_filter,
    })




from django.shortcuts import render, get_object_or_404
from .models import Ningen_Staff, CustomUser, Activite

def agent_detail(request, id_beside):

    ningen_staff = get_object_or_404(Ningen_Staff, ID_Beside=id_beside)


    custom_user = get_object_or_404(CustomUser, name=ningen_staff.ID_Beside)

    activities = Activite.objects.filter(agent=ningen_staff)

    context = {
        'custom_user': custom_user,
        'ningen_staff': ningen_staff,
        'activities': activities,
    }

    return render(request, 'agent_detail.html', context)



import pandas as pd

from django.http import HttpResponse
from django.shortcuts import render
from .models import Ningen_Staff, Activite

def download_selected_staff(request):
    if request.method == 'POST':
        staff_ids = request.POST.getlist('selected_staff[]')
        selected_staff = Ningen_Staff.objects.filter(ID_Beside__in=staff_ids)

        staff_data = []
        for staff in selected_staff:
            staff_info = {
                "ID_Beside": staff.ID_Beside,
                "Prenom": staff.Prenom,
                "Nom": staff.Nom,
                "Ngroupe": staff.Ngroupe,
                "Matricule": staff.Matricule,
                "Fonction":staff.Fonction,
                "Genre":staff.Genre,
                "CL": ", ".join([cl.name for cl in staff.CL.all()]),                
                "N1": staff.N1,
                "N2": staff.N2,
                "Date_embauche": staff.Date_embauche,
                "Date_fin_1ere_période_d_essai": staff.Date_fin_1ere_période_d_essai,
                "Type_de_contrat": staff.Type_de_contrat,

                "Régime_horaire": staff.Régime_horaire,
                "CIN": staff.CIN,
                "Titulaire": staff.Titulaire,
                "Date_de_titularisation": staff.Date_de_titularisation,
                "Tel": staff.Tel,
                "Date_de_naissance": staff.Date_de_naissance,
                "age": staff.age,
                "Situation": staff.Situation,
                "Nombre_d_enfants": staff.Nombre_d_enfants,
                "Niveau_d_etude": staff.Niveau_d_etude,
                "Diplome": staff.Diplome,
                "Centre_d_intérêt": staff.Centre_d_intérêt,

                "Mail_personel": staff.Mail_personel,
                "Adresse_postale": staff.Adresse_postale,
                "Activité_actuelle": staff.Activité_actuelle,
                "Micro_activité_actuelle": staff.Micro_activité_actuelle,
                "LOB_actuelle": staff.LOB_actuelle,
                "Affectation_actuelle_CANAL": staff.Affectation_actuelle_CANAL,
                "TT": staff.TT,
                "Statut_à_date": staff.Statut_à_date,
                "Date_FPE": staff.Date_FPE,
                "Mission_spéciale": staff.Mission_spéciale,
                }

            # Get related Activite data for the staff member
            activite_data = Activite.objects.filter(agent=staff)

            # Iterate through related Activite objects and add their data to the dictionary
            for activite in activite_data:
                activite_info = {
                    f"Activité {activite.numero_activite} - Nom": activite.nom_activite,
                    f"Activité {activite.numero_activite} - Line of business": activite.Line_of_business,
                    f"Activité {activite.numero_activite} - Skills_CS": activite.Skills_CS,
                    f"Activité {activite.numero_activite} - Skills_Sales": activite.Skills_Sales,
                    f"Activité {activite.numero_activite} - Skills_BO": activite.Skills_BO,
                    f"Activité {activite.numero_activite} - Skills_Confirmation": activite.Skills_Confirmation,
                    f"Activité {activite.numero_activite} - CANAL_IB": activite.CANAL_IB,
                    f"Activité {activite.numero_activite} - CANAL_OB": activite.CANAL_OB,
                    f"Activité {activite.numero_activite} - CANAL_Live_Chat": activite.CANAL_Live_Chat,
                }
                staff_info.update(activite_info)

            staff_data.append(staff_info)

        df = pd.DataFrame(staff_data)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=selected_staff.xlsx'

        df.to_excel(response, index=False)

        return response

    return render(request, 'staff_list.html')



from django.shortcuts import render, get_object_or_404, redirect
from .models import Ningen_Staff

def modify_staff(request, id_beside):
    staff_member = get_object_or_404(Ningen_Staff, ID_Beside=id_beside)

    if request.method == 'POST':
        form = NingenStaffForm(request.POST, instance=staff_member)
        if form.is_valid():
            form.save()
            return redirect('agent_detail', id_beside=id_beside)
    else:
        form = NingenStaffForm(instance=staff_member)

    return render(request, 'modify_staff.html', {'form': form, 'staff_member': staff_member})


def modify_activity(request):
    agent = request.GET.get('agent')
    numero_activite = request.GET.get('numero_activite')

    activity = get_object_or_404(Activite, agent__ID_Beside=agent, numero_activite=numero_activite)
    form = ActiviteForm(instance=activity)

    if request.method == 'POST':
        form = ActiviteForm(request.POST, instance=activity)
        if form.is_valid():
            form.save()
            return redirect('agent_detail', id_beside=agent)
    
    return render(request, 'modify_activity.html', {'form': form, 'activity': activity})




from django.shortcuts import get_object_or_404

def add_activity(request):
    agent_name = request.GET.get('agent')
    print(agent_name)
    if request.method == 'POST':
        form = ActiviteForm(request.POST)
        if form.is_valid():
            # Get the agent object based on the name in the URL
            agent = get_object_or_404(Ningen_Staff, ID_Beside=agent_name)

            # Assign the agent to the form's instance
            form.instance.agent = agent

            # Save the activity
            new_activity = form.save()
            return redirect('agent_detail', id_beside=agent_name)
    else:
        # Pass the agent to the form when rendering
        form = ActiviteForm(initial={'agent': agent_name})

    return render(request, 'add_activity.html', {'form': form, 'agent': agent_name})



def create_management(request):
    if request.method == 'POST':
        form = NingenManagementForm(request.POST)
        if form.is_valid():
            Ningen_Management = form.save()
            return redirect('index')
    else:
        form = NingenManagementForm()

    return render(request, 'addManagement.html', {'form': form})



def testview(request):
    return render(request, 'test.html')



from django.http import HttpResponse
# Assurez-vous d'importer la classe ExcelUploadForm correctement
from .forms import ExcelUploadForm  # Assurez-vous d'adapter le chemin d'importation selon votre structure de répertoire
import pandas as pd
from datetime import datetime, timedelta

def delete_workhours_by_id_upload(id_upload):
    # Utilisez la méthode filter pour obtenir un ensemble de requêtes filtré
    workhours_to_delete = WorkHours.objects.filter(id_upload=id_upload)

    # Utilisez la méthode delete pour supprimer les objets filtrés
    workhours_to_delete.delete()
@login_required
def WorkHour(request):
    print("trraitement des heures à afficher  ...")
    # Récupérez les données de la base de données
    current_user = request.user

    # Filtrez les données de la base de données pour correspondre à l'utilisateur
    work_hours = WorkHours.objects.filter(full_name__iexact=current_user.name).order_by('date')


    # Créez une liste d'événements au format JSON adapté à FullCalendar
    events = []
    previous_workhour = None
    more_previous_workhour = None
    hourpasoff_debut = None
    hourpasoff_fin = None
    for workhour in work_hours:
        if workhour.traitement == "OFF" :
            if hourpasoff_fin != None :
                if int(hourpasoff_fin[:-3]) < 10:
                    print("haw el date",str(workhour.date + timedelta(days=1)) + " 0" + str(int(hourpasoff_fin[:-3]) + 1) + ":00")
                    event = {
                        'title': workhour.traitement,
                        'start': str(workhour.date) + " " + str(hourpasoff_debut),  # Format de date et heure adapté
                        'end': str(workhour.date + timedelta(days=1)) + " 0" + str(int(hourpasoff_fin[:-3]) + 1) + ":00",
                        'color': 'gray',
                    }
                else :
                    event = {
                        'title': workhour.traitement,
                        'start': str(workhour.date) + " " + str(hourpasoff_debut),  # Format de date et heure adapté
                        'end': str(workhour.date) + " " + str(int(hourpasoff_fin[:-3]) + 1) +":00",
                        'color': 'gray',
                    }
            else :
                event = {
                    'title': workhour.traitement,
                    'start': str(workhour.date) + " " + "09:00",  # Format de date et heure adapté
                    'end': str(workhour.date) + " " + "19:00",
                    'color': 'gray',
                }
        else :
            if workhour.traitement == "CP":
                event = {
                    'title': "Congé payé",
                    'start': str(workhour.date) + " " + "09:00",  # Format de date et heure adapté
                    'end': str(workhour.date) + " " + "24:00",
                    'color': '#eb75b6',
                }
            else :
                if workhour.traitement == "fin":
                    hourpasoff_fin = workhour.hour
                    if int(workhour.hour[:-3]) < 10 :
                        event = {
                            'title': workhour.traitement,
                            'start': str(workhour.date) + " "+str(workhour.hour),
                            'color': 'red',
                        }
                    else:
                        event = {
                            'title': workhour.traitement,
                            'start': str(workhour.date) + " " + str(workhour.hour),  # Format de date et heure adapté
                            'color': 'red',
                        }
                else :
                    if workhour.traitement == "debut" :
                        hourpasoff_debut = workhour.hour
                        event = {
                            'title': workhour.traitement,
                            'start': str(workhour.date) + " " + str(workhour.hour),  # Format de date et heure adapté
                            'color': 'green',
                        }
                    else :
                        if workhour.traitement == "Pause_dej":
                            if int(workhour.hour[3:]) != 00 :
                                event = {
                                    'title': previous_workhour.traitement,
                                    'start': str(workhour.date) + " " + str(int(previous_workhour.hour[:-3]) + 1) +":00",
                                    'end': str(workhour.date) + " " + str(workhour.hour),
                                }

                                events.append(event)
                                finworkhour = datetime.strptime(workhour.hour, "%H:%M") + timedelta(hours=1)
                                event = {
                                    'title': workhour.traitement,
                                    'start': str(workhour.date) + " " + str(workhour.hour),
                                    'end': str(workhour.date) + " " + finworkhour.strftime("%H:%M"),
                                    'color': 'orange',
                                }
                            else :
                                event = {
                                    'title': workhour.traitement,
                                    'start': str(workhour.date) + " " + str(workhour.hour),  # Format de date et heure adapté
                                    'end': str(workhour.date) + " " + str(int(workhour.hour[:-3]) + 1) +":00",
                                    'color': 'orange',
                                }
                        else :
                            if previous_workhour.traitement == "Pause_dej" and previous_workhour.hour != 00 :
                                finworkhour = datetime.strptime(previous_workhour.hour, "%H:%M") + timedelta(hours=1)
                                if int(workhour.hour[:-3]) < 10:
                                    event = {
                                        'title': workhour.traitement,
                                        'start': str(workhour.date) + " " + finworkhour.strftime("%H:%M"),
                                        'end': str(workhour.date) + " 0" + str(int(workhour.hour[:-3]) + 1) + ":00",
                                    }
                                else :
                                    event = {
                                        'title': workhour.traitement,
                                        'start': str(workhour.date) + " " + finworkhour.strftime("%H:%M"),
                                        'end': str(workhour.date) + " " + str(int(workhour.hour[:-3])+1)+":00",
                                    }
                            else :
                                if int(workhour.hour[:-3]) < 10:
                                    event = {
                                        'title': workhour.traitement,
                                        'start': str(workhour.date) + " " + str(workhour.hour),
                                        'end': str(workhour.date) + " 0" + str(int(workhour.hour[:-3]) + 1) + ":00",
                                    }
                                else :
                                    event = {
                                        'title': workhour.traitement,
                                        'start': str(workhour.date) + " " +str(workhour.hour),
                                        'end': str(workhour.date) + " " + str(int(workhour.hour[:-3])+1)+":00",
                                    }
        more_previous_workhour=previous_workhour
        previous_workhour = workhour
        events.append(event)
    # Passez les données JSON à votre modèle HTML
    context = {
        'events': events,
    }
    return render(request, 'WorkHour.html', context)
def fonction_upload_planing(xl,uploada):

    global idsupr,predone
    print("debut de lupload ...")
    dernier_id_jour = WorkHours.objects.all().order_by('-id_day').first()
    dernier_id_upload = upload.objects.all().order_by('-id_upload').first()

    if dernier_id_jour:
        dernier_id_jour = dernier_id_jour.id_day
    else:
        dernier_id_jour = 0

    if dernier_id_upload:
        dernier_id_upload = dernier_id_upload.id_upload
    else:
        dernier_id_upload = 0

    ligne = 7
    full_name_suiv = " "
    id_day = int(dernier_id_jour) + 1
    id_upload = int(dernier_id_upload) + 1
    predone = int(dernier_id_upload)
    idsupr = id_upload

    while pd.isnull(full_name_suiv) == False:
        # Utiliser .loc[] pour accéder à la cellule
        full_name = xl.iloc[ligne, 0]
        full_name_suiv = xl.iloc[ligne + 1, 0]
        print(full_name)
        print("name suiv ", full_name_suiv)
        traitement = xl.iloc[ligne, 1]
        print("traitement ",traitement)
        place = xl.iloc[ligne, 2]
        print("place",place)
        print("dernier_id_jour",dernier_id_jour)
        for i in range(0, 35, 5):
            print(i)
            print("id_day",id_day)
            heure_debut = xl.iloc[ligne, i + 3]
            date = str(xl.iloc[5, i + 3])[:-9]
            print(date)
            print(heure_debut)
            heure_fin = str(xl.iloc[ligne, i + 6])
            print(heure_fin)
            if (pd.isnull(heure_debut) == False) and (heure_debut != "CP"):
                heure_debut = str(xl.iloc[ligne, i + 3])[:-3]
                print(heure_debut)
                heure_fin = str(xl.iloc[ligne, i + 6])[:-3]
                if len(heure_fin)>7 :
                    heure_fin = str(xl.iloc[ligne, i + 6])[:-3]
                    heure_fin = heure_fin[11:]

                print("haaw el fin",heure_fin)
                if pd.isnull(xl.iloc[ligne, i + 4]):
                    dej = False
                else :
                    dej = True
                heure_dej = str(xl.iloc[ligne, i + 4])[:-3]
                print(heure_dej)
                print(heure_fin)
                un_jour = timedelta(days=1)
                done = datetime.strptime(heure_fin, '%H:%M')
                print(done)
                compteur = datetime.strptime(heure_debut, '%H:%M')
                hours = datetime.strptime("00:00", '%H:%M')
                while compteur != done + un_jour:
                    hours = hours + timedelta(hours=1)
                    compteur = compteur + timedelta(hours=1)
                    print(compteur, done + un_jour)

                print("heure_debut :", heure_debut)
                print("heure_fin :", heure_fin)
                hours = int(datetime.strftime(hours, '%H:%M')[:-3])
                print("hours :", hours)
                heured = datetime.strptime(heure_debut, '%H:%M')

                while hours >= 0:
                    if dej == True :
                        h_dej = int(heure_dej[:-3])
                    else :
                        h_dej = 2555
                    if int(datetime.strftime(heured, '%H:%M')[:-3]) != h_dej:
                        print("comparaison heure et heured", heure_debut, heure_fin,
                              datetime.strftime(heured, '%H:%M'))
                        if heure_debut == datetime.strftime(heured, '%H:%M'):
                            work_hour = WorkHours(
                                full_name=full_name,
                                traitement="debut",
                                hour=datetime.strftime(heured, '%H:%M'),
                                date=date,
                                place=place,
                                id_day = id_day,
                                id_upload = id_upload
                            )

                        else:
                            print("heure ya sehbi :",int(datetime.strftime(heured, '%H:%M')[:-3]))
                            if 6 > int(datetime.strftime(heured, '%H:%M')[:-3]):
                                date = datetime.strptime(date, '%Y-%m-%d')
                                work_hour = WorkHours(
                                    full_name=full_name,
                                    traitement=traitement,
                                    hour=datetime.strftime(heured, '%H:%M'),
                                    date=str(date + timedelta(days=1))[:-9],
                                    place=place,
                                    id_day=id_day,
                                    id_upload=id_upload
                                )
                                date = str(date)[:-9]
                                print(date)
                            else :
                                work_hour = WorkHours(
                                    full_name=full_name,
                                    traitement=traitement,
                                    hour=datetime.strftime(heured, '%H:%M'),
                                    date=date,
                                    place=place,
                                    id_day = id_day,
                                    id_upload = id_upload
                                )
                    else:
                        work_hour = WorkHours(
                            full_name=full_name,
                            traitement="Pause_dej",
                            hour=heure_dej,
                            date=date,
                            place=place,
                            id_day = id_day,
                            id_upload=id_upload
                        )

                    print(heure_fin ,datetime.strftime(heured, '%H:%M'))
                    if heure_fin == datetime.strftime(heured, '%H:%M'):

                        if 6 > int(datetime.strftime(heured, '%H:%M')[:-3]):
                            date = datetime.strptime(date, '%Y-%m-%d')
                            work_hour = WorkHours(
                                full_name=full_name,
                                traitement="fin",
                                hour=datetime.strftime(heured, '%H:%M'),
                                date=str(date + timedelta(days=1))[:-9],
                                place=place,
                                id_day=id_day,
                                id_upload=id_upload
                            )
                            date = str(date)[:-9]
                        else:
                            work_hour = WorkHours(
                                full_name=full_name,
                                traitement="fin",
                                hour=datetime.strftime(heured, '%H:%M'),
                                date=date,
                                place=place,
                                id_day=id_day,
                                id_upload=id_upload
                            )
                    print(hours)
                    print("jenvoie ca à la base de données : ", work_hour.full_name, work_hour.traitement,
                          work_hour.hour, work_hour.date, work_hour.place,work_hour.id_day,work_hour.id_upload)
                    if uploada == True :
                        work_hour.save()
                    heured = heured + timedelta(hours=1)
                    hours = hours - 1

            else:
                if heure_debut == "CP":
                    work_hour = WorkHours(
                        full_name=full_name,
                        traitement="CP",
                        hour="--",
                        date=date,
                        place=place,
                        id_day=id_day,
                        id_upload=id_upload
                    )
                    if uploada == True:
                        work_hour.save()
                    print("conje payé")
                if pd.isnull(heure_debut) == True:
                    work_hour = WorkHours (
                        full_name=full_name,
                        traitement="OFF",
                        hour="--",
                        date=date,
                        place=place,
                        id_day=id_day,
                        id_upload=id_upload
                    )
                    if uploada == True:
                        work_hour.save()
                    print("OFF")
                    done = int(id_upload)
                    print("done ----------------------------- : ",done)
            id_day = id_day + 1
        print(
            "\n------------------------------------------------------------------N E X T------------------------------------------------------------------------\n")
        ligne = ligne + 1

    print('dernier id_day :',dernier_id_jour)

from datetime import date
@login_required
def upload_planning(request):
    if request.method == 'POST':
        print("debut de lupload ...")
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            if excel_file.name.endswith('.xlsx'):
                xl = pd.read_excel(excel_file, engine='openpyxl')
                try :
                    fonction_upload_planing(xl,uploada=False)
                except Exception as e :
                    print("Une erreur s'est produite :", e)
                    return 'Le tableau des plannings doit remplir les critères'
                else :
                    try :
                        fonction_upload_planing(xl,uploada=True)
                        upload(id_upload=idsupr,nom_excel=excel_file,date=date.today()).save()
                        print("data_base filled ! ")
                        return 'Planning Enregistré'
                    except Exception as e :
                        print("Une erreur s'est produite :", e)
                        if idsupr != predone :
                            print("je supprime cet id :",idsupr)
                            delete_workhours_by_id_upload(idsupr)
                        return 'Doublon(s) détecté(s) , Verifiez vos données puis Réessayez'

            else:
                return 'Le fichier doit être au format Excel (.xlsx)'

from django.shortcuts import render, redirect, get_object_or_404
from .models import WorkHours
from .forms import WorkHourForm
from .forms import WorkHourFilterForm

global work_hourssession
affichersession = 1

@login_required
def manage_hours(request):

    class Session:
        def __init__(self, fullname, date, traitement, debut, pause, fin):
            self.fullname = fullname
            self.date = date
            self.traitement = traitement
            self.debut = debut
            self.pause = pause
            self.fin = fin

        def to_dict(self):
            return {
                'fullname': self.fullname,
                'date': self.date,
                'traitement': self.traitement,
                'debut': self.debut,
                'pause': self.pause,
                'fin': self.fin,
            }

    if affichersession != 0 :
        session_json = request.session.get('session_data', '{}')

        # Convertissez le JSON en dictionnaire
        session_data = json.loads(session_json)

        # Récupérez les valeurs nécessaires
        work_hourssession_data = session_data.get('work_hourssession', [])
        nom_excel_data = session_data.get('nom_excel', '')
        print(nom_excel_data)
        # Créez des objets Session
        work_hourssession = [Session(**data) for data in work_hourssession_data]
    else :
        work_hourssession = None
        nom_excel_data = None

    erorupload = upload_planning(request)
    formplanning = ExcelUploadForm()

    if request.method == 'POST':
        form = WorkHourForm(request.POST)
        if form.is_valid():
            form.save()
    else:
        form = WorkHourForm()

    filter_form = WorkHourFilterForm(request.GET)  # Utilisez les paramètres GET pour initialiser le formulaire

    for upload_instance in upload.objects.all():
        count = WorkHours.objects.filter(id_upload=upload_instance.id_upload).count()
        upload_instance.workhour_dispo = count
        upload_instance.save()
    uploads = upload.objects.filter(workhour_dispo__gt=0)
    uploads = uploads.order_by('-date')

    return render(request, 'upload_planning.html', {'nom_excel': nom_excel_data[:-5],'work_hours': work_hourssession, 'form': formplanning, 'filter_form': filter_form, 'erorupload': erorupload,'uploads':uploads})

def delete_work_hour(request, work_hour_id):
    work_hour = get_object_or_404(WorkHours, pk=work_hour_id)
    work_hour.delete()
    return redirect('upload_planning')

@login_required
def delete_upload(request, upload_id):

    try:
        delete_workhours_by_id_upload(upload_id)
        return redirect('upload_planning')
    except Exception as e:

        return redirect('upload_planning')

def show_session(request, upload_id):

    work_hourssession = []
    class Session:
        def __init__(self, fullname, date, traitement, debut, pause, fin):
            self.fullname = fullname
            self.date = date.strftime('%Y-%m-%d') if date else "-"
            self.traitement = traitement
            self.debut = debut
            self.pause = pause
            self.fin = fin

        def to_dict(self):
            return {
                'fullname': self.fullname,
                'date': self.date,
                'traitement': self.traitement,
                'debut': self.debut,
                'pause': self.pause,
                'fin': self.fin,
            }


    distinct_id_days_count = WorkHours.objects.filter(id_upload=upload_id).values('id_day').distinct()

    for i in distinct_id_days_count:
        id_day = i['id_day']

        if WorkHours.objects.filter(id_upload=upload_id,id_day=id_day,traitement="OFF").values_list('full_name', flat=True) :
            fullname_session_qs = WorkHours.objects.filter(id_upload=upload_id,id_day=id_day,traitement="OFF").values_list('full_name', flat=True)
            date_session_qs = WorkHours.objects.filter(id_upload=upload_id,id_day=id_day,traitement="OFF").values_list('date', flat=True)
            debut_session_qs = WorkHours.objects.filter(id_upload=upload_id,id_day=id_day,traitement="OFF").values_list('traitement', flat=True)
            traitement_session_qs = WorkHours.objects.filter(id_upload=upload_id,id_day=id_day,traitement="OFF").values_list('traitement', flat=True)
            pause_session_qs = WorkHours.objects.filter(id_upload=upload_id,id_day=id_day,traitement="OFF").values_list('traitement', flat=True)
            fin_session_qs = WorkHours.objects.filter(id_upload=upload_id,id_day=id_day,traitement="OFF").values_list('traitement', flat=True)
        else :
            if WorkHours.objects.filter(id_upload=upload_id,id_day=id_day,traitement="CP").values_list('full_name', flat=True):
                id = WorkHours.objects.filter(id_upload=upload_id,id_day=id_day, traitement="CP").values_list('id', flat=True)[0]
                fullname_session_qs = WorkHours.objects.filter(id_upload=upload_id,id_day=id_day, traitement="CP").values_list('full_name', flat=True)
                date_session_qs = WorkHours.objects.filter(id_upload=upload_id,id_day=id_day, traitement="CP").values_list('date',flat=True)
                traitement_session_qs = WorkHours.objects.filter(id_upload=upload_id,id_day=id_day, traitement="CP").values_list('traitement',flat=True)
                debut_session_qs = WorkHours.objects.filter(id_upload=upload_id,id_day=id_day, traitement="CP").values_list('traitement',flat=True)
                pause_session_qs = WorkHours.objects.filter(id_upload=upload_id,id_day=id_day, traitement="CP").values_list('traitement',flat=True)
                fin_session_qs = WorkHours.objects.filter(id_upload=upload_id,id_day=id_day, traitement="CP").values_list('traitement',flat=True)
            else :

                id = WorkHours.objects.filter(id_upload=upload_id,id_day=id_day,traitement="debut").values_list('id', flat=True)[0]
                fullname_session_qs = WorkHours.objects.filter(id_upload=upload_id,id_day=id_day,traitement="debut").values_list('full_name', flat=True)
                date_session_qs = WorkHours.objects.filter(id_upload=upload_id,id_day=id_day, traitement="debut").values_list('date', flat=True)
                traitement_session_qs = WorkHours.objects.filter(id_upload=upload_id,id_day=id_day,id=id+1).values_list('traitement', flat=True)
                debut_session_qs = WorkHours.objects.filter(id_upload=upload_id,id_day=id_day,traitement="debut").values_list('hour', flat=True)
                pause_session_qs = WorkHours.objects.filter(id_upload=upload_id,id_day=id_day,traitement="Pause_dej").values_list('hour',flat=True)
                fin_session_qs = WorkHours.objects.filter(id_upload=upload_id,id_day=id_day,traitement="fin").values_list('hour', flat=True)

        fullname_session = fullname_session_qs[0] if fullname_session_qs else "-"
        date_session = date_session_qs[0] if date_session_qs else "-"
        traitement_session = traitement_session_qs[0] if traitement_session_qs else "-"
        debut_session = debut_session_qs[0] if debut_session_qs else "-"
        pause_session = pause_session_qs[0] if pause_session_qs else "-"
        fin_session = fin_session_qs[0] if fin_session_qs else "-"

        work_hourssession.append(Session(fullname_session,date_session,traitement_session,debut_session, pause_session, fin_session))

    work_hourssession_data = [session.to_dict() for session in work_hourssession]
    nom_excel_data = upload.objects.filter(id_upload=upload_id).values_list('nom_excel', flat=True).first()

    session_data = {
        'work_hourssession': work_hourssession_data,
        'nom_excel': nom_excel_data,
    }
    session_json = json.dumps(session_data)
    request.session['session_data'] = session_json

    return redirect('upload_planning')